import pandas as pd
from openpyxl import load_workbook

texto = """

"""

nomes = [ 

    "Gessica Lima",
    "Gilyane Souza",
    "Joao Santana",
    "Julia Costa",
    "Julie Moreira",
    "KLEYTON JESUS",
    "Lucas Silva",
    "Thaissa Santos",
    
    "JACKSON YAMAMOTO",
    "Dilceia Machado",
    "Tiago Menezes",
    "INC"
]
print ("Equipe help desk")
contagens = {nome: texto.count(nome) for nome in nomes}

for nome, contagem in contagens.items():
    print(f'-"{nome}"  {contagem} Chamados')

inc_nome = "INC"
contagem_inc = contagens.get(inc_nome, 0)

total_sem_inc = sum(contagem for nome, contagem in contagens.items() if nome != inc_nome)

resultado_final = abs(total_sem_inc - contagem_inc)

print("  |___Total de chamados")

print(f"\nFila com um total de {resultado_final} Chamados não designados.")

df = pd.DataFrame(list(contagens.items()), columns=["Nome", "Contagem"])

caminho_arquivo = "QuantidadeChamados.xlsx"
book = load_workbook(caminho_arquivo)


nome_aba = "Resultados"
if nome_aba in book.sheetnames:
    with pd.ExcelWriter(caminho_arquivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=nome_aba, index=False)
else:
    with pd.ExcelWriter(caminho_arquivo, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name=nome_aba, index=False)

print(f"\nResultados exportados para 'QuantidadeChamados.xlsx' na aba '{nome_aba}'.")
